import os
import logging
from typing import Dict, Any, Optional

from openpyxl import load_workbook

from .config import INSTRUMENTS, TIMEZONE
from .trade_time import get_trade_context
from .time_utils import _now_tz
from .db_client import (
    ensure_schema,
    is_initialized,
    mark_initialized,
    clear_initialized_flag,
    reset_all_data,
    drop_and_recreate_schema,
    backfill_strong_signals,
    upsert_instrument,
    insert_instrument_data,
    get_last_strong_signal,
    set_position_state,
)

logger = logging.getLogger(__name__)

DEFAULT_XLSX_PATH = "/data/форекс.xlsx"


def _env_init_mode() -> str:
    """
    DB_INIT_MODE:
      - RESET   : жёсткий сброс (очистка старой БД), затем импорт из Excel, затем восстановление сигналов
      - MIGRATE : обновление схемы без потери данных (migrations), импорт из Excel только если БД ещё не инициализирована
      - NONE    : ничего с БД не делаем
    """
    # Backward-compat:
    if os.getenv("RESET_DB_ON_BOOT", "0").strip() == "1":
        return "RESET"

    mode = os.getenv("DB_INIT_MODE", "MIGRATE").strip().upper()
    if mode in ("0", "OFF", "NO", "FALSE", "NONE", ""):
        return "NONE"
    if mode not in ("RESET", "MIGRATE", "NONE"):
        logger.warning("Неизвестный DB_INIT_MODE=%r, использую MIGRATE.", mode)
        return "MIGRATE"
    return mode


def main() -> None:
    mode = _env_init_mode()
    logger.info("DB init mode: %s", mode)

    if mode == "NONE":
        logger.info("DB_INIT_MODE=NONE — пропускаю любую инициализацию/миграции.")
        return

    # 1) RESET or MIGRATE schema
    if mode == "RESET":
        strategy = os.getenv("DB_RESET_STRATEGY", "DROP").strip().upper()
        logger.warning("DB_INIT_MODE=RESET — выполняю сброс БД (strategy=%s).", strategy)

        if strategy == "TRUNCATE":
            # Обновляем схему, затем чистим данные
            ensure_schema()
            reset_all_data()
            # После TRUNCATE схемы уже есть, но на всякий случай прогоняем migrations ещё раз
            ensure_schema()
        else:
            # DROP+recreate — самый надёжный вариант для смены схемы
            drop_and_recreate_schema()

        # На reset всегда снимаем флаг initialized, чтобы импорт точно выполнился
        clear_initialized_flag()

    else:
        seen = {}
        for instr in INSTRUMENTS:
            if instr.sheet_name in seen and seen[instr.sheet_name] != instr.code:
                raise RuntimeError(
                    f"Дубликат sheet_name='{instr.sheet_name}' для {seen[instr.sheet_name]} и {instr.code}")
            seen[instr.sheet_name] = instr.code

        # MIGRATE
        ensure_schema()

        # ВАЖНО: всегда регистрируем/обновляем инструменты в instruments
        for instr in INSTRUMENTS:
            upsert_instrument(
                code=instr.code,
                sheet_name=instr.sheet_name,
                plus500_name=instr.plus500_name,
                signals_enabled=bool(getattr(instr, "signals_enabled", False)),
            )

        # 2) Если уже инициализировано — импорт пропускаем (но instruments уже актуальны)
        if mode != "RESET" and is_initialized():
            logger.info("БД уже инициализирована — импорт из Excel пропускаю.")

            if os.getenv("DB_BACKFILL_SIGNALS", "0").strip() == "1":
                inserted = backfill_strong_signals()
                logger.info("Backfill strong signals: inserted=%d", inserted)

            return

    # 3) Импорт из Excel
    xlsx_path = os.getenv("FOREX_XLSX_PATH", DEFAULT_XLSX_PATH)
    logger.info("Импорт в БД из Excel: %s", xlsx_path)

    if not os.path.exists(xlsx_path):
        logger.error("Файл Excel не найден: %s", xlsx_path)
        return

    wb = load_workbook(xlsx_path, data_only=True)

    # 3.1) Регистрируем в БД все инструменты из INSTRUMENTS
    for instr in INSTRUMENTS:
        upsert_instrument(
            code=instr.code,
            sheet_name=instr.sheet_name,
            plus500_name=instr.plus500_name,
            signals_enabled=bool(getattr(instr, "signals_enabled", False)),
        )

    total_data = 0
    total_skipped = 0

    # 3.2) Импортируем строки из каждого листа
    for instr in INSTRUMENTS:
        if instr.sheet_name not in wb.sheetnames:
            logger.warning("В Excel нет листа %s для инструмента %s — пропускаю.", instr.sheet_name, instr.code)
            continue

        ws = wb[instr.sheet_name]

        # Ожидаемый формат: строки с временем/ценой/таймфреймами.
        # Ваш текущий парсер уже знает как прочитать нужные колонки: оставляем как есть.
        # Ниже — безопасный проход: пропуск пустых строк.
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            # В вашем файле обычно: (datetime, price, 30m, 1h, 5h, 1d, 1w, 1mo) — но порядок может отличаться.
            # Если у вас другой формат, вы уже обрабатывали его ранее — этот блок можно адаптировать.
            try:
                candle_time = row[0]
                price = row[1]
                tf_30m = row[2] if len(row) > 2 else None
                tf_1h = row[3] if len(row) > 3 else None
                tf_5h = row[4] if len(row) > 4 else None
                tf_1d = row[5] if len(row) > 5 else None
                tf_1w = row[6] if len(row) > 6 else None
                tf_1mo = row[7] if len(row) > 7 else None

                if not candle_time:
                    total_skipped += 1
                    continue

                # Как было у вас: если любой таймфрейм пустой — пропускаем строку
                if any(v is None for v in (tf_30m, tf_1h, tf_5h, tf_1d, tf_1w, tf_1mo)):
                    total_skipped += 1
                    continue

                # Нормализуем timeframes в формат, который ожидает insert_instrument_data()
                timeframes: Dict[str, Any] = {
                    "30m": {"action": tf_30m},
                    "1h": {"action": tf_1h},
                    "5h": {"action": tf_5h},
                    "1d": {"action": tf_1d},
                    "1w": {"action": tf_1w},
                    "1mo": {"action": tf_1mo},
                }

                ok = insert_instrument_data(
                    instrument_code=instr.code,
                    executed_at=str(candle_time),
                    price=float(price or 0.0),
                    timeframes=timeframes,
                )
                if ok:
                    total_data += 1
                else:
                    total_skipped += 1

            except Exception:
                total_skipped += 1
                logger.exception("Ошибка импорта строки Excel для %s: %r", instr.code, row)

    # 4) Восстанавливаем instrument_signals из instrument_data (безопасно, можно включать всегда)
    if os.getenv("DB_BACKFILL_SIGNALS", "1").strip() == "1":
        inserted = backfill_strong_signals()
        logger.info("Backfill strong signals: inserted=%d", inserted)

    # 5) Bootstrap position_state: только для инструментов, которые СЕЙЧАС в TRADE/CLOSE_ONLY.
    now_dt = _now_tz(TIMEZONE)
    for instr in INSTRUMENTS:
        ctx = get_trade_context(instr_name=instr.code, sheet_name=instr.sheet_name, tz_name=TIMEZONE, dt=now_dt)
        if ctx.mode not in ("TRADE", "CLOSE_ONLY"):
            continue

        tf_key = getattr(ctx, "timeframe_key", "30m")
        last_sig = get_last_strong_signal(instr.code, timeframe=str(tf_key))

        if last_sig == "STRONG BUY":
            set_position_state(instr.code, "LONG", comment=f"bootstrap: tf={tf_key}, last_sig={last_sig}")
        elif last_sig == "STRONG SELL":
            set_position_state(instr.code, "SHORT", comment=f"bootstrap: tf={tf_key}, last_sig={last_sig}")
        else:
            set_position_state(instr.code, "NONE", comment=f"bootstrap: tf={tf_key}, last_sig=None")

    mark_initialized()
    logger.info(
        "Инициализация завершена. Данные: %d, пропущено строк: %d",
        total_data, total_skipped
    )


if __name__ == "__main__":
    main()